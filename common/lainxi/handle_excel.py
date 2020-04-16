"""
北京-漠雪殇-27
"""
import openpyxl

class handel_excel:
    def __init__(self,filename,sheel):
        self.filename = filename
        self.sheel = sheel

    def read_excel(self):
        lw = openpyxl.load_workbook(self.filename)
        sh = lw[self.sheel]
        re = list(sh.rows)
        case_data = []
        title = []
        for tit in re[0]:
            title.append(tit.value)

        for n in re[1:]:
            li = []
            for i in n:
                li.append(i.value)
                cas = dict(zip(title,li))
                case_data.append(cas)
        return case_data

    def write_excel(self,row,column,value):
        lw = openpyxl.load_workbook(self.filename)
        sh = lw[self.sheel]
        sh.cell(row=row,column=column,value=value)
        lw.save(self.filename)





