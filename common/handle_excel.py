"""
北京-漠雪殇-27
"""
import openpyxl

class Excel_Hande:
    def __init__(self,filename,sheet):
        """

        :param filename: excel文件路径
        :param sheet: 表单
        """
        self.filename = filename
        self.sheet = sheet
    def read_excel(self):
        lw = openpyxl.load_workbook(self.filename)
        sh = lw[self.sheet]
        res = list(sh.rows)
        case_data = []
        tit = []
        for i in res[0]:
            tit.append(i.value)

        for item in res[1:]:
            values = []
            for n in item:
                values.append(n.value)
            dic = dict(zip(tit,values))
            case_data.append(dic)
        return case_data
    def write(self,row,column,value):
        """

        :param row: 行
        :param column:列
        :param value: 值
        :return: 一个格子值
        """
        lw = openpyxl.load_workbook(self.filename)
        sh = lw[self.sheet]
        sh.cell(row=row,column=column,value=value)
        lw.save(self.filename)





if __name__ == '__main__':
    from pprint import pprint
    exce = Excel_Hande(r'E:\pycharm_new\api_test\data\apicases.xlsx','register')
    k = exce.read_excel()
    pprint(type(k[0]))





